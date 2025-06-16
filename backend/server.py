from fastapi import FastAPI, APIRouter, HTTPException, File, UploadFile
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime
import json
import zipfile
import io
import base64
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
from openpyxl import Workbook
import barcode
from barcode.writer import ImageWriter, SVGWriter
import qrcode
from PIL import Image
import tempfile
import shutil


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Data Models
class CustomerDetails(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    surname: str
    organization: str
    country: str
    address: str
    phone: str
    email: str
    gst_number: Optional[str] = None
    state: Optional[str] = None

class BarcodeOrder(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_details: CustomerDetails
    barcode_type: str
    quantity: int
    total_amount: float
    tax_amount: float
    final_amount: float
    order_status: str = "pending"  # pending, processing, completed, failed
    payment_status: str = "pending"  # pending, paid, failed
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class BarcodeOrderCreate(BaseModel):
    customer_details: CustomerDetails
    barcode_type: str
    quantity: int

class PaymentDetails(BaseModel):
    order_id: str
    payment_method: str  # razorpay, paypal
    transaction_id: Optional[str] = None
    payment_status: str = "pending"

# Barcode Configuration - Prices in Indian Rupees (INR)
BARCODE_TYPES = {
    "qr_code": {"name": "QR Code", "price": 150.0},  # ₹150 per barcode
    "code128": {"name": "Code 128", "price": 120.0},  # ₹120 per barcode
    "ean13": {"name": "EAN-13", "price": 140.0},  # ₹140 per barcode
    "upc": {"name": "UPC-A", "price": 140.0},  # ₹140 per barcode
    "code39": {"name": "Code 39", "price": 120.0},  # ₹120 per barcode
    "datamatrix": {"name": "Data Matrix", "price": 180.0}  # ₹180 per barcode
}

# Tax Configuration for India
TAX_RATES = {
    "gujarat": {"cgst": 9.0, "sgst": 9.0, "total": 18.0},
    "other": {"igst": 18.0, "total": 18.0}
}

# Utility Functions
def calculate_tax_and_total(base_amount: float, state: str) -> Dict[str, float]:
    """Calculate tax and total amount based on state (for India)"""
    if state.lower() == "gujarat":
        tax_amount = base_amount * 0.18
        return {
            "base_amount": base_amount,
            "tax_amount": tax_amount,
            "cgst": base_amount * 0.09,
            "sgst": base_amount * 0.09,
            "total_amount": base_amount + tax_amount
        }
    else:
        tax_amount = base_amount * 0.18
        return {
            "base_amount": base_amount,
            "tax_amount": tax_amount,
            "igst": tax_amount,
            "total_amount": base_amount + tax_amount
        }

def generate_barcode_data(barcode_type: str, quantity: int) -> List[Dict]:
    """Generate barcode data based on type and quantity"""
    barcodes = []
    for i in range(quantity):
        barcode_id = f"{barcode_type.upper()}{str(uuid.uuid4())[:8]}"
        barcodes.append({
            "id": barcode_id,
            "type": barcode_type,
            "data": barcode_id,
            "generated_at": datetime.utcnow().isoformat()
        })
    return barcodes

def create_barcode_image(barcode_data: str, barcode_type: str) -> str:
    """Generate barcode image and return as base64"""
    try:
        if barcode_type == "qr_code":
            qr = qrcode.QRCode(version=1, box_size=10, border=5)
            qr.add_data(barcode_data)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            
            # Convert to base64
            buffer = io.BytesIO()
            img.save(buffer, format='PNG')
            img_str = base64.b64encode(buffer.getvalue()).decode()
            return img_str
        else:
            # For other barcode types, use python-barcode
            code_class = barcode.get_barcode_class(barcode_type)
            code = code_class(barcode_data, writer=ImageWriter())
            
            buffer = io.BytesIO()
            code.write(buffer)
            img_str = base64.b64encode(buffer.getvalue()).decode()
            return img_str
    except Exception as e:
        print(f"Error generating barcode: {e}")
        return ""

def create_invoice_data(order: BarcodeOrder, tax_details: Dict) -> Dict:
    """Create invoice data structure with INR currency"""
    return {
        "order_id": order.id,
        "customer": order.customer_details.dict(),
        "items": [{
            "description": f"{BARCODE_TYPES[order.barcode_type]['name']} Barcode",
            "quantity": order.quantity,
            "unit_price": BARCODE_TYPES[order.barcode_type]['price'],
            "total": order.quantity * BARCODE_TYPES[order.barcode_type]['price']
        }],
        "tax_details": tax_details,
        "total_amount": order.final_amount,
        "currency": "INR",
        "date": order.created_at.strftime("%Y-%m-%d")
    }

# API Routes
@api_router.get("/")
async def root():
    return {"message": "Barcode Generation Service API"}

@api_router.get("/barcode-types")
async def get_barcode_types():
    """Get available barcode types and their prices in INR"""
    return {"barcode_types": BARCODE_TYPES, "currency": "INR"}

@api_router.post("/calculate-price")
async def calculate_price(barcode_type: str, quantity: int, state: str = "other"):
    """Calculate price including tax for given barcode type and quantity (in INR)"""
    if barcode_type not in BARCODE_TYPES:
        raise HTTPException(status_code=400, detail="Invalid barcode type")
    
    if quantity <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be greater than 0")
    
    base_price = BARCODE_TYPES[barcode_type]["price"]
    base_amount = base_price * quantity
    
    tax_details = calculate_tax_and_total(base_amount, state)
    
    return {
        "barcode_type": barcode_type,
        "quantity": quantity,
        "unit_price": base_price,
        "pricing": tax_details,
        "currency": "INR"
    }

@api_router.post("/create-order")
async def create_order(order_data: BarcodeOrderCreate):
    """Create a new barcode order with INR pricing"""
    try:
        # Validate barcode type
        if order_data.barcode_type not in BARCODE_TYPES:
            raise HTTPException(status_code=400, detail="Invalid barcode type")
        
        # Calculate pricing in INR
        base_price = BARCODE_TYPES[order_data.barcode_type]["price"]
        base_amount = base_price * order_data.quantity
        
        state = order_data.customer_details.state or "other"
        tax_details = calculate_tax_and_total(base_amount, state)
        
        # Create order
        order = BarcodeOrder(
            customer_details=order_data.customer_details,
            barcode_type=order_data.barcode_type,
            quantity=order_data.quantity,
            total_amount=base_amount,
            tax_amount=tax_details["tax_amount"],
            final_amount=tax_details["total_amount"]
        )
        
        # Save to database
        await db.barcode_orders.insert_one(order.dict())
        
        return {
            "order_id": order.id,
            "order": order.dict(),
            "tax_details": tax_details,
            "currency": "INR",
            "message": "Order created successfully"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to create order: {str(e)}")

@api_router.get("/order/{order_id}")
async def get_order(order_id: str):
    """Get order details by ID"""
    order = await db.barcode_orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Convert MongoDB ObjectId to string to make it JSON serializable
    if "_id" in order:
        order["_id"] = str(order["_id"])
    
    return order

@api_router.post("/process-order/{order_id}")
async def process_order(order_id: str):
    """Process order - generate barcodes and create zip file"""
    try:
        # Get order from database
        order_data = await db.barcode_orders.find_one({"id": order_id})
        if not order_data:
            raise HTTPException(status_code=404, detail="Order not found")
        
        # Convert MongoDB ObjectId to string
        if "_id" in order_data:
            order_data["_id"] = str(order_data["_id"])
        
        order = BarcodeOrder(**order_data)
        
        # Update order status to processing
        await db.barcode_orders.update_one(
            {"id": order_id},
            {"$set": {"order_status": "processing", "updated_at": datetime.utcnow()}}
        )
        
        # Generate barcode data
        barcode_list = generate_barcode_data(order.barcode_type, order.quantity)
        
        # Create in-memory zip file
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Create Excel file with barcode data
            wb = Workbook()
            ws = wb.active
            ws.title = "Barcode_Data"
            
            # Headers
            ws['A1'] = "Barcode ID"
            ws['B1'] = "Type"
            ws['C1'] = "Data"
            ws['D1'] = "Generated At"
            
            # Add barcode data
            for idx, bc in enumerate(barcode_list, 2):
                ws[f'A{idx}'] = bc['id']
                ws[f'B{idx}'] = bc['type']
                ws[f'C{idx}'] = bc['data']
                ws[f'D{idx}'] = bc['generated_at']
            
            # Save Excel to buffer
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            # Add Excel file to zip
            zip_file.writestr("barcode_data.xlsx", excel_buffer.getvalue())
            
            # Generate and add barcode images
            for idx, bc in enumerate(barcode_list):
                img_base64 = create_barcode_image(bc['data'], bc['type'])
                if img_base64:
                    img_data = base64.b64decode(img_base64)
                    zip_file.writestr(f"barcodes/{bc['id']}.png", img_data)
            
            # Create invoice data with INR
            state = order.customer_details.state or "other"
            tax_details = calculate_tax_and_total(order.total_amount, state)
            invoice_data = create_invoice_data(order, tax_details)
            
            # Add invoice JSON
            zip_file.writestr("invoice.json", json.dumps(invoice_data, indent=2))
        
        # Update order status to completed
        await db.barcode_orders.update_one(
            {"id": order_id},
            {"$set": {"order_status": "completed", "updated_at": datetime.utcnow()}}
        )
        
        # Return zip file
        zip_buffer.seek(0)
        
        return StreamingResponse(
            io.BytesIO(zip_buffer.read()),
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename=barcodes_{order_id}.zip"}
        )
        
    except Exception as e:
        # Update order status to failed
        await db.barcode_orders.update_one(
            {"id": order_id},
            {"$set": {"order_status": "failed", "updated_at": datetime.utcnow()}}
        )
        raise HTTPException(status_code=500, detail=f"Failed to process order: {str(e)}")

@api_router.get("/orders")
async def list_orders(limit: int = 50):
    """List all orders with pagination"""
    orders = await db.barcode_orders.find().limit(limit).to_list(limit)
    
    # Convert MongoDB ObjectId to string for JSON serialization
    for order in orders:
        if "_id" in order:
            order["_id"] = str(order["_id"])
    
    return {"orders": orders}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()